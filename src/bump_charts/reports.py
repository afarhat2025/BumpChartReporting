"""Report writing and email delivery."""

import logging
import os
import smtplib
from email.message import EmailMessage
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from .config import (
    SMTP_SERVER, SMTP_PORT, SMTP_USER, SUCCESS_RECIPIENTS, 
    ERROR_RECIPIENT, EMAIL_SUBJECT, RESULT_COLUMNS
)
from .models import PriceResult

logger = logging.getLogger("compare_prices")


def write_results(results: List[PriceResult], output_filepath: str, send_email: bool = True) -> None:
    """
    Write price comparison results to Excel and email.
    
    Args:
        results: List of PriceResult objects
        output_filepath: Path to output Excel file
    """
    if not results:
        logger.warning("No results to write.")
        return
    
    # Convert results to DataFrame
    data = [result.to_dict() for result in results]
    df_results = pd.DataFrame(data)[RESULT_COLUMNS]
    
    try:
        # Write Excel file
        _write_excel_file(df_results, output_filepath)
        
        # Send emails (optional)
        if send_email:
            _send_success_emails([output_filepath])
    
    except PermissionError as perm_exc:
        error_msg = f"Could not write to {output_filepath}. Please close the file if it is open and try again."
        logger.error(error_msg)
        _send_error_email(error_msg)
    
    except Exception as exc:
        error_msg = f"Error writing {output_filepath}: {str(exc)}"
        logger.exception(error_msg)
        _send_error_email(error_msg)


def _write_excel_file(df_results: pd.DataFrame, output_filepath: str) -> None:
    """Write results to Excel file with formatting."""
    output_dir = os.path.dirname(output_filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    
    if os.path.exists(output_filepath):
        try:
            os.remove(output_filepath)
        except PermissionError:
            logger.warning(
                "Existing report %s is locked; attempting to overwrite anyway.",
                output_filepath
            )
    
    # Write to Excel
    df_results.to_excel(output_filepath, index=False)
    
    # Format the workbook
    wb = load_workbook(output_filepath)
    ws = wb.active
    ws.freeze_panes = ws["A2"]
    
    # Hide Part_Key column
    part_key_col_idx = len(RESULT_COLUMNS)
    col_letter = ws.cell(row=1, column=part_key_col_idx).column_letter
    ws.column_dimensions[col_letter].hidden = True
    
    wb.save(output_filepath)
    logger.info("Report written to %s (Part_Key column hidden)", output_filepath)


def send_success_email_with_attachments(attachment_paths: List[str]) -> None:
    """Public helper to send one success email with multiple attachments."""
    _send_success_emails(attachment_paths)


def _send_success_emails(attachment_paths: List[str]) -> None:
    """Send success emails with one or more report attachments."""
    for recipient in SUCCESS_RECIPIENTS:
        try:
            _send_email(
                recipient,
                EMAIL_SUBJECT,
                "Price comparison report attached.",
                attachment_paths
            )
            logger.info(
                "Email sent to %s with attachments %s",
                recipient,
                ", ".join([os.path.basename(p) for p in attachment_paths if p])
            )
        except Exception as email_exc:
            logger.exception("Failed to send email to %s", recipient)


def _send_error_email(error_msg: str) -> None:
    """Send error notification email."""
    try:
        _send_email(
            ERROR_RECIPIENT,
            f"{EMAIL_SUBJECT} - ERROR",
            error_msg,
            None
        )
        logger.info("Error notification sent to %s", ERROR_RECIPIENT)
    except Exception as email_exc:
        logger.exception("Failed to send error email to %s", ERROR_RECIPIENT)


def _send_email(
    to_email: str,
    subject: str,
    body: str,
    attachment_paths: Optional[List[str]] = None
) -> None:
    """
    Send email with optional one or many attachments.
    
    Args:
        to_email: Recipient email address
        subject: Email subject
        body: Email body
        attachment_paths: Optional list of attachment file paths
    """
    msg = EmailMessage()
    msg["From"] = SMTP_USER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)
    
    if attachment_paths:
        import mimetypes
        for attachment_path in attachment_paths:
            if attachment_path and os.path.exists(attachment_path):
                mimetype = mimetypes.guess_type(attachment_path)[0] or "application/octet-stream"
                maintype, subtype = mimetype.split("/", 1)
                with open(attachment_path, "rb") as f:
                    file_data = f.read()
                    msg.add_attachment(
                        file_data,
                        maintype=maintype,
                        subtype=subtype,
                        filename=os.path.basename(attachment_path)
                    )
    
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.send_message(msg)
